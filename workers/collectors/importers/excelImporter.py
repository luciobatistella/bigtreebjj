import json
import sys
from typing import Any, Dict

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover
    openpyxl = None


def import_excel(file_path: str, options: Dict[str, Any] | None = None) -> Dict[str, Any]:
    options = options or {}
    worksheet_name = options.get("worksheet") or None
    if openpyxl is None:
        return {
            "format": "xlsx",
            "file_name": file_path.split('/')[-1],
            "rows": [],
            "summary": {"total_rows": 0, "valid_rows": 0, "invalid_rows": 0},
            "preview": "[]",
            "metadata": {"worksheet": worksheet_name},
        }

    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[worksheet_name] if worksheet_name and worksheet_name in workbook.sheetnames else workbook.active
    rows = []
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=2):
        rows.append({"row_number": index, "raw": list(row)})
    return {
        "format": "xlsx",
        "file_name": file_path.split('/')[-1],
        "rows": rows,
        "summary": {"total_rows": len(rows), "valid_rows": len(rows), "invalid_rows": 0},
        "preview": json.dumps(rows[:3], ensure_ascii=False),
        "metadata": {"worksheet": sheet.title},
    }


if __name__ == "__main__":
    options = json.loads(sys.argv[2]) if len(sys.argv) > 2 else {}
    print(json.dumps(import_excel(sys.argv[1], options)))
